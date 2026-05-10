"""
md_to_excel.py
==============
Convierte RESULTADOS_TESTS.md a un archivo Excel con:
  - Una hoja por cada sección de datos (CSV embebidos + tablas MD)
  - Estilos: cabeceras en azul oscuro, filas alternas, bordes, columnas ajustadas
  - Hoja de resumen ejecutivo al inicio

Uso:
    python scripts/md_to_excel.py
Salida:
    RESULTADOS_TESTS.xlsx  (en la raíz del proyecto)
"""

import os
import re
import csv
import io

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Rutas ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MD_FILE  = os.path.join(BASE_DIR, "RESULTADOS_TESTS.md")
OUT_FILE = os.path.join(BASE_DIR, "RESULTADOS_TESTS.xlsx")

# ── Paleta de estilos ─────────────────────────────────────────────────────────
HDR_FILL    = PatternFill("solid", fgColor="1F4E79")   # azul oscuro
HDR_FONT    = Font(color="FFFFFF", bold=True, size=10)
EVEN_FILL   = PatternFill("solid", fgColor="D6E4F0")   # azul muy claro
ODD_FILL    = PatternFill("solid", fgColor="FFFFFF")
PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")   # verde
FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")   # rojo
TITLE_FONT  = Font(bold=True, size=12, color="1F4E79")
CELL_FONT   = Font(size=9)
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

SECTION2_FILL = PatternFill("solid", fgColor="2E75B6")  # azul medio (sección 8)

def _hdr_fill_for(section_idx):
    return HDR_FILL if section_idx < 8 else SECTION2_FILL


# ── Helpers ───────────────────────────────────────────────────────────────────

def style_header_row(ws, row_num, n_cols, section_idx=0):
    fill = _hdr_fill_for(section_idx)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_data_row(ws, row_num, n_cols, even=True):
    fill = EVEN_FILL if even else ODD_FILL
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER


def auto_width(ws, min_w=8, max_w=45):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, min_w), max_w)


def write_csv_sheet(wb, sheet_name, csv_text, section_idx=0):
    reader = csv.reader(io.StringIO(csv_text.strip()))
    rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        return
    ws = wb.create_sheet(title=sheet_name[:31])
    headers = rows[0]
    ws.append(headers)
    style_header_row(ws, 1, len(headers), section_idx)
    for i, row in enumerate(rows[1:], start=2):
        ws.append(row)
        style_data_row(ws, i, len(headers), even=(i % 2 == 0))
        # Colorear columnas PASS/FAIL
        for col_idx, (hdr, val) in enumerate(zip(headers, row), start=1):
            if hdr.lower() == "fail" and val.strip() not in ("", "0"):
                ws.cell(row=i, column=col_idx).fill = FAIL_FILL
            if hdr.lower() in ("pass", "tasa_exito") and val.strip() not in ("", "0"):
                ws.cell(row=i, column=col_idx).fill = PASS_FILL
    ws.freeze_panes = "A2"
    auto_width(ws)
    return ws


def write_md_table_sheet(wb, sheet_name, md_table_text, section_idx=0):
    """Convierte una tabla Markdown |col|col| a hoja Excel."""
    lines = [l.strip() for l in md_table_text.strip().splitlines()
             if l.strip() and not re.match(r"^\|[-| :]+\|$", l.strip())]
    if not lines:
        return
    def parse_row(line):
        return [c.strip() for c in line.strip("|").split("|")]

    ws = wb.create_sheet(title=sheet_name[:31])
    headers = parse_row(lines[0])
    ws.append(headers)
    style_header_row(ws, 1, len(headers), section_idx)
    for i, line in enumerate(lines[1:], start=2):
        row = parse_row(line)
        ws.append(row)
        style_data_row(ws, i, len(headers), even=(i % 2 == 0))
        for col_idx, (hdr, val) in enumerate(zip(headers, row), start=1):
            v = val.strip()
            if hdr.lower() == "fallaron" and v not in ("", "0"):
                ws.cell(row=i, column=col_idx).fill = FAIL_FILL
            if hdr.lower() in ("pasaron", "pass", "tasa") and v not in ("", "0"):
                ws.cell(row=i, column=col_idx).fill = PASS_FILL
    ws.freeze_panes = "A2"
    auto_width(ws)
    return ws


# ── Parser del Markdown ───────────────────────────────────────────────────────

def extract_sections(md_text):
    """
    Devuelve lista de (titulo_seccion, tipo, contenido):
      tipo = 'csv' | 'mdtable'
    """
    sections = []

    # CSV blocks: ```csv ... ```
    csv_pattern = re.compile(
        r"###\s+(.+?)\n.*?```csv\n(.*?)```",
        re.DOTALL
    )
    for m in csv_pattern.finditer(md_text):
        title = m.group(1).strip()
        content = m.group(2)
        sections.append((title, "csv", content))

    # MD tables after ### headings (not already captured)
    table_pattern = re.compile(
        r"###\s+(.+?)\n((?:\|.+\n)+)",
        re.MULTILINE
    )
    captured_titles = {s[0] for s in sections}
    for m in table_pattern.finditer(md_text):
        title = m.group(1).strip()
        if title in captured_titles:
            continue
        content = m.group(2)
        if content.count("|") > 2:
            sections.append((title, "mdtable", content))
            captured_titles.add(title)

    # Top-level ## sections with MD tables
    top_pattern = re.compile(
        r"##\s+(?!\d+\.\d+)(.+?)\n((?:(?!\n##)[\s\S])*?)(?=\n##|\Z)",
        re.MULTILINE
    )
    for m in top_pattern.finditer(md_text):
        title = m.group(1).strip()
        body  = m.group(2)
        table_m = re.search(r"((?:\|.+\n)+)", body)
        if table_m and title not in captured_titles:
            sections.append((title, "mdtable", table_m.group(1)))
            captured_titles.add(title)

    return sections


def get_section_idx(title):
    """Determina si pertenece al bloque de benchmark (sección 8)."""
    return 8 if re.search(r"8\.", title) or "Benchmark" in title or "Endpoint" in title or "Ranking" in title or "STDEV" in title or "AVG vs" in title or "MIN / AVG" in title else 0


# ── Hoja de Resumen ───────────────────────────────────────────────────────────

def build_summary_sheet(wb, md_text):
    ws = wb.create_sheet(title="RESUMEN", index=0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22

    title_cell = ws["A1"]
    title_cell.value = "Análisis de Resultados — API Ambulancias SOAT"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30

    sub = ws["A2"]
    sub.value = "28 de abril de 2026  |  Contenedor Docker backend"
    sub.font = Font(italic=True, size=9, color="595959")
    ws.merge_cells("A2:B2")
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 18

    kpis = [
        ("── SUITE DE TESTS ──", None),
        ("Total de tests",        "146"),
        ("Pasaron (OK)",          "146"),
        ("Fallaron (FAIL)",       "0"),
        ("Tasa de éxito",         "100 %"),
        ("Tiempo de ejecución",   "18.870 s"),
        ("── BENCHMARK ──", None),
        ("Endpoints medidos",     "31"),
        ("Total peticiones",      "620"),
        ("AVG global respuesta",  "2.70 ms"),
        ("Endpoint más rápido",   "Sedes GET detalle — 0.98 ms"),
        ("Endpoint más lento",    "Reporte Excel — 6.10 ms (sostenido)"),
        ("POST más lento",        "Sedes POST crear — 5.69 ms"),
    ]

    row = 4
    for label, value in kpis:
        if value is None:
            # separador / subtítulo
            c = ws.cell(row=row, column=1, value=label)
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = HDR_FILL
            c.alignment = LEFT
            c.border = BORDER
            ws.merge_cells(f"A{row}:B{row}")
        else:
            cl = ws.cell(row=row, column=1, value=label)
            cv = ws.cell(row=row, column=2, value=value)
            fill = EVEN_FILL if row % 2 == 0 else ODD_FILL
            for c in (cl, cv):
                c.fill = fill
                c.font = CELL_FONT
                c.border = BORDER
            cl.alignment = LEFT
            cv.alignment = CENTER
            # Colorear valor 0 / 100%
            if value in ("0", "0 %"):
                cv.fill = PASS_FILL
            if value == "100 %":
                cv.fill = PASS_FILL
        row += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    with open(MD_FILE, encoding="utf-8") as f:
        md_text = f.read()

    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    # 1. Hoja de resumen
    build_summary_sheet(wb, md_text)

    # 2. Hojas de datos
    sections = extract_sections(md_text)
    seen = {}
    for title, kind, content in sections:
        # Nombre único de hoja (máx 31 chars)
        safe = re.sub(r"[\\/*?:\[\]]", "", title)[:28]
        if safe in seen:
            seen[safe] += 1
            safe = f"{safe[:26]}_{seen[safe]}"
        else:
            seen[safe] = 0

        sidx = get_section_idx(title)

        if kind == "csv":
            write_csv_sheet(wb, safe, content, sidx)
        else:
            write_md_table_sheet(wb, safe, content, sidx)

    wb.save(OUT_FILE)
    print(f"Excel generado: {OUT_FILE}")
    print(f"Hojas creadas : {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
