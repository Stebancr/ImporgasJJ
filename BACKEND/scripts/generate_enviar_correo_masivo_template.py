import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

TEMPLATE_HEADERS = [
    "Empresa",
    "Unidad",
    "Proyecto",
    "Centro",
    "Nombre",
    "CC",
    "Ciudad",
    "Cargo",
    "TipoExamen",
    "Examenes",
]

SAMPLE_ROWS = [
    ["EMPRESA ABC", "UNIDAD NORTE", "PROYECTO 1", "CENTRO 101", "Juan Pérez", "123456789", "Bogotá", "Operario", "INGRESO", "Examen Médico, Prueba Psicológica"],
    ["EMPRESA XYZ", "UNIDAD SUR", "PROYECTO 2", "CENTRO 202", "María Gómez", "987654321", "Medellín", "Supervisor", "PERIODICO", "Audiometría, Espirometría"],
]


def generate_template(output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "EnviarCorreoMasivo"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    for col_num, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Sample rows
    for r_idx, row in enumerate(SAMPLE_ROWS, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Save file
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    target = os.path.join("examenes", "template", "EnviarCorreoMasivo_template.xlsx")
    path = os.path.abspath(target)
    generate_template(path)
    print(f"Template generated at: {path}")
