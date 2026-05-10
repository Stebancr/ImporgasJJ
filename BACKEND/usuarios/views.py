import json
import csv
import io
import os
from io import BytesIO
from django.http import JsonResponse, FileResponse
from django.db import transaction
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from usuarios.permissions import IsSuperAdmin, IsAdminUser, IsUsuarioEspecial, IsSuperUserOrAdmin
from usuarios.models import Usuario, Credenciales, Cargo, Niveles, Regional
from usuarios.serializers import UsuarioListadoSerializer, CargoSerializer, NivelesSerializer, RegionalSerializer
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


class Perfil(APIView):
    """Vista de perfil del usuario."""
    permission_classes = [IsAuthenticated]

    def get(self, request, id=None):
        if id is not None:
            usuario = Usuario.objects.select_related('cargo', 'nivel', 'regional').filter(id=id).first()
        else:
            usuario_rel = getattr(request.user, 'usuario_rel', None)
            if usuario_rel:
                usuario = Usuario.objects.select_related('cargo', 'nivel', 'regional').filter(id=usuario_rel.id).first()
            else:
                usuario = None

        if not usuario:
            return Response({"error": "El usuario no tiene perfil asociado"}, status=400)

        data = {
            "id": usuario.id,
            "cedula": usuario.cedula,
            "nombre_completo": usuario.nombre_completo,
            "correo": usuario.correo,
            "telefono": usuario.telefono,
            "sede": usuario.sede,
            "estado": usuario.estado,
            "nombre_cargo": usuario.cargo.nombrecargo if usuario.cargo else None,
            "nombre_nivel": usuario.nivel.nombrenivel if usuario.nivel else None,
            "nombre_regional": usuario.regional.nombreregional if usuario.regional else None,
        }
        return Response(data)

    def patch(self, request, id=None):
        """Alterna el estado del usuario (0 <-> 1). Requiere id."""
        if id is None:
            return Response({"error": "Se requiere el id del usuario"}, status=400)

        usuario = Usuario.objects.filter(id=id).first()
        if not usuario:
            return Response({"error": "Usuario no encontrado"}, status=404)

        nuevo_estado = 0 if usuario.estado == 1 else 1
        usuario.estado = nuevo_estado
        usuario.save()

        credencial = getattr(usuario, 'credenciales', None)
        if credencial:
            credencial.estado = nuevo_estado
            credencial.save()

        return Response({
            "id": usuario.id,
            "nuevo_estado": nuevo_estado,
        })


class Register(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsAdminUser]

    def post(self, request, *args, **kwargs):
        payload = request.data if hasattr(request, 'data') else None
        if not payload:
            try:
                payload = json.loads(request.body.decode('utf-8'))
            except Exception:
                return JsonResponse({'error': 'JSON invalido'}, status=400)

        required = ['usuario', 'password', 'cedula', 'nombre_completo']
        missing = [k for k in required if (isinstance(payload.get(k), str) and not payload.get(k, '').strip()) or (not isinstance(payload.get(k), str) and k not in payload)]
        if missing:
            return JsonResponse({'error': f'Faltan campos requeridos: {missing}'}, status=400)

        usuario_nombre = payload.get('usuario', '').strip()
        if Credenciales.objects.filter(usuario=usuario_nombre).exists():
            return JsonResponse({'error': f'El usuario "{usuario_nombre}" ya existe'}, status=400)

        cedula = str(payload.get('cedula', '')).strip()
        if Usuario.objects.filter(cedula=cedula).exists():
            return JsonResponse({'error': f'El usuario con cedula "{cedula}" ya existe'}, status=400)

        try:
            with transaction.atomic():
                usuario_obj = Usuario.objects.create(
                    cedula=cedula,
                    nombre_completo=payload.get('nombre_completo', '').strip(),
                    correo=payload.get('correo') or '',
                    telefono=payload.get('telefono') or '',
                    sede=payload.get('sede') or None,
                    cargo_id=payload.get('cargo') or None,
                    nivel_id=payload.get('nivel') or None,
                    regional_id=payload.get('regional') or None,
                    estado=1,
                )

                cred = Credenciales(
                    usuario=usuario_nombre,
                    tipo_usuario=int(payload.get('tipo_usuario', 0)),
                    usuario_rel=usuario_obj,
                    estado=1,
                )
                cred.set_password(payload['password'])
                cred.save()

            return JsonResponse({
                'mensaje': 'Usuario creado correctamente',
                'usuario_id': cred.id,
                'usuario_rel_id': usuario_obj.id,
            }, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    def get(self, request, usuario_id):
        usuario = (
            Usuario.objects
            .select_related('cargo', 'nivel', 'regional')
            .filter(id=usuario_id)
            .first()
        )
        if not usuario:
            return Response({"error": "Usuario no encontrado"}, status=404)

        data = {
            "id": usuario.id,
            "cedula": usuario.cedula,
            "nombre_completo": usuario.nombre_completo,
            "correo": usuario.correo,
            "telefono": usuario.telefono,
            "sede": usuario.sede,
            "cargo": usuario.cargo_id,
            "nivel": usuario.nivel_id,
            "regional": usuario.regional_id,
            "estado": usuario.estado,
        }
        return Response(data)

    def put(self, request, usuario_id):
        usuario = Usuario.objects.filter(id=usuario_id).first()
        if not usuario:
            return Response({"error": "Usuario no encontrado"}, status=404)

        data = request.data
        mapeo = {
            'nombre_completo': 'nombre_completo',
            'correo': 'correo',
            'telefono': 'telefono',
            'sede': 'sede',
            'cargo': 'cargo_id',
            'nivel': 'nivel_id',
            'regional': 'regional_id',
        }
        for campo_front, campo_modelo in mapeo.items():
            if campo_front in data:
                setattr(usuario, campo_modelo, data[campo_front])
        usuario.save()

        serializer = UsuarioListadoSerializer(usuario)
        return Response(serializer.data)


class RegisterTemporal(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsUsuarioEspecial]

    def post(self, request, *args, **kwargs):
        payload = request.data if hasattr(request, 'data') else None
        if not payload:
            try:
                payload = json.loads(request.body.decode('utf-8'))
            except Exception:
                return JsonResponse({'error': 'JSON invalido'}, status=400)

        required = ['usuario', 'password', 'cedula', 'nombre_completo']
        if any(key not in payload for key in required):
            return JsonResponse({'error': f'Faltan campos requeridos: {required}'}, status=400)

        usuario_nombre = payload.get('usuario', '').strip()
        if not usuario_nombre:
            return JsonResponse({'error': 'El usuario no puede estar vacio'}, status=400)
        if Credenciales.objects.filter(usuario=usuario_nombre).exists():
            return JsonResponse({'error': f'El usuario {usuario_nombre} ya existe'}, status=400)

        cedula = str(payload.get('cedula', '')).strip()
        if not cedula:
            return JsonResponse({'error': 'La cedula no puede estar vacia'}, status=400)
        if Usuario.objects.filter(cedula=cedula).exists():
            return JsonResponse({'error': f'La cedula {cedula} ya existe'}, status=400)

        try:
            with transaction.atomic():
                usuario_obj = Usuario.objects.create(
                    cedula=cedula,
                    nombre_completo=payload.get('nombre_completo', '').strip(),
                    correo=payload.get('correo') or '',
                    telefono=payload.get('telefono') or '',
                    sede=payload.get('sede') or None,
                    cargo_id=payload.get('cargo') or None,
                    nivel_id=payload.get('nivel') or None,
                    regional_id=payload.get('regional') or None,
                    estado=1,
                )

                cred = Credenciales(
                    usuario=usuario_nombre,
                    tipo_usuario=0,
                    usuario_rel=usuario_obj,
                    estado=1,
                )
                cred.set_password(payload['password'])
                cred.save()

            return JsonResponse({
                'mensaje': 'Usuario temporal creado',
                'usuario_id': cred.id,
                'usuario_rel_id': usuario_obj.id,
            }, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class ListaUsuarios(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsAdminUser]

    def get(self, request, *args, **kwargs):
        try:
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            if page < 1:
                page = 1
            if page_size < 1 or page_size > 100:
                page_size = 10

            search = request.GET.get('search', '').strip()

            base_qs = (
                Usuario.objects
                .filter(estado=1)
                .select_related('cargo', 'nivel', 'regional')
                .order_by('id')
            )

            if search:
                base_qs = base_qs.filter(
                    Q(nombre_completo__icontains=search) |
                    Q(cedula__icontains=search) |
                    Q(correo__icontains=search)
                )

            total = base_qs.count()
            start = (page - 1) * page_size
            end = start + page_size
            items = list(base_qs[start:end])

            results = UsuarioListadoSerializer(items, many=True).data

            return Response({
                'count': total,
                'page': page,
                'page_size': page_size,
                'results': results,
            })
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class CargoNivelRegionalView(APIView):
    """Vista para obtener listas de Cargo, Niveles y Regionales."""
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperUserOrAdmin]

    def get(self, request):
        cargos = Cargo.objects.filter(estadocargo=1).order_by('nombrecargo')
        niveles = Niveles.objects.filter(estadonivel=1).order_by('nombrenivel')
        regionales = Regional.objects.filter(estadoregional=1).order_by('nombreregional')

        return Response({
            "cargos": CargoSerializer(cargos, many=True).data,
            "niveles": NivelesSerializer(niveles, many=True).data,
            "regionales": RegionalSerializer(regionales, many=True).data,
        })


class FiltrarUsuariosView(APIView):
    """Vista para filtrar usuarios por nombre o cedula."""
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsAdminUser]

    def get(self, request):
        query = request.GET.get('q', '').strip()
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        if page < 1:
            page = 1
        if page_size < 1 or page_size > 100:
            page_size = 10

        base_qs = (
            Usuario.objects
            .exclude(estado=3)
            .select_related('cargo', 'nivel', 'regional')
            .order_by('id')
        )
        if query:
            base_qs = base_qs.filter(
                Q(nombre_completo__icontains=query) |
                Q(cedula__icontains=query) |
                Q(correo__icontains=query)
            )

        total = base_qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        items = list(base_qs[start:end])
        results = UsuarioListadoSerializer(items, many=True).data

        return Response({
            'count': total,
            'page': page,
            'page_size': page_size,
            'results': results,
        })


class CambiarEstadoUsuarioView(APIView):
    """Vista para activar o desactivar usuarios (uno o multiples)."""
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    def patch(self, request, usuario_id=None):
        """PATCH individual: /usuarios/cambiar-estado-usuario/<usuario_id>/"""
        try:
            if usuario_id is None:
                return Response({"error": "usuario_id es requerido en la URL"}, status=400)

            nuevo_estado = request.data.get('estado')
            if nuevo_estado is None:
                return Response({"error": "El campo 'estado' es requerido (0 o 1)"}, status=400)
            if nuevo_estado not in [0, 1]:
                return Response({"error": "El estado debe ser 0 (inactivo) o 1 (activo)"}, status=400)

            usuario = Usuario.objects.filter(id=usuario_id).first()
            if not usuario:
                return Response({"error": "Usuario no encontrado"}, status=404)

            usuario.estado = nuevo_estado
            usuario.save()

            credencial = getattr(usuario, 'credenciales', None)
            if credencial:
                credencial.estado = nuevo_estado
                credencial.save()

            return Response({
                "mensaje": "Estado actualizado correctamente",
                "usuario_id": usuario.id,
                "nuevo_estado": nuevo_estado,
            }, status=200)

        except Exception as e:
            return Response({"error": f"Error al actualizar estado: {str(e)}"}, status=500)

    def post(self, request):
        """POST masivo: /usuarios/cambiar-estado-usuario/"""
        try:
            usuario_ids = request.data.get('usuario_ids', [])
            cedulas = request.data.get('cedulas', [])
            nuevo_estado = request.data.get('estado')

            if len(usuario_ids) == 0 and len(cedulas) == 0:
                return Response({"error": "'usuario_ids' o 'cedulas' requerido y no vacio"}, status=400)

            if nuevo_estado is None:
                return Response({"error": "El campo 'estado' es requerido (0 o 1)"}, status=400)
            if nuevo_estado not in [0, 1]:
                return Response({"error": "El estado debe ser 0 (inactivo) o 1 (activo)"}, status=400)

            encontrados = []
            no_encontrados = []
            actualizados = 0

            ids_a_buscar = list(usuario_ids)
            for cedula in cedulas:
                u = Usuario.objects.filter(cedula=str(cedula)).first()
                if u:
                    ids_a_buscar.append(u.id)
                else:
                    no_encontrados.append({"identificador": str(cedula), "tipo": "cedula", "error": "Cedula no encontrada"})

            for uid in ids_a_buscar:
                try:
                    usuario = Usuario.objects.filter(id=uid).first()
                    if not usuario:
                        no_encontrados.append({"identificador": uid, "tipo": "usuario_id", "error": "Usuario no encontrado"})
                        continue

                    usuario.estado = nuevo_estado
                    usuario.save()

                    credencial = getattr(usuario, 'credenciales', None)
                    if credencial:
                        credencial.estado = nuevo_estado
                        credencial.save()

                    encontrados.append({
                        "usuario_id": usuario.id,
                        "cedula": usuario.cedula,
                        "nombre": usuario.nombre_completo,
                        "estado": nuevo_estado,
                        "success": True,
                    })
                    actualizados += 1
                except Exception as e:
                    no_encontrados.append({"identificador": uid, "tipo": "usuario_id", "error": str(e)})

            return Response({
                "mensaje": f"Procesamiento completado: {actualizados} actualizados, {len(no_encontrados)} no encontrados",
                "total": len(usuario_ids) + len(cedulas),
                "actualizados": actualizados,
                "no_encontrados": len(no_encontrados),
                "detalles_encontrados": encontrados,
                "detalles_no_encontrados": no_encontrados,
            }, status=200)

        except Exception as e:
            return Response({"error": f"Error al procesar estados: {str(e)}"}, status=500)


class ActualizarRolUsuarioView(APIView):
    """Vista para cambiar el rol de un usuario. Solo SuperAdmin."""
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    def patch(self, request, usuario_id):
        try:
            nuevo_rol = request.data.get('tipo_usuario')
            if nuevo_rol is None:
                return Response({"error": "El campo 'tipo_usuario' es requerido"}, status=400)
            if nuevo_rol not in [0, 1, 2, 3, 4]:
                return Response({"error": "El tipo_usuario debe ser 0, 1, 2, 3 o 4"}, status=400)

            usuario = Usuario.objects.filter(id=usuario_id).first()
            if not usuario:
                return Response({"error": "Usuario no encontrado"}, status=404)

            credencial = getattr(usuario, 'credenciales', None)
            if not credencial:
                return Response({"error": "El usuario no tiene credenciales asociadas"}, status=404)

            rol_anterior = credencial.tipo_usuario
            credencial.tipo_usuario = nuevo_rol
            credencial.save()

            return Response({
                "mensaje": "Rol actualizado correctamente",
                "usuario_id": usuario.id,
                "rol_anterior": rol_anterior,
                "nuevo_rol": nuevo_rol,
            }, status=200)

        except Exception as e:
            return Response({"error": f"Error al actualizar rol: {str(e)}"}, status=500)


class DatosCargoView(APIView):
    """CRUD para gestionar Cargos."""
    permission_classes = [IsAuthenticated, IsSuperUserOrAdmin | IsUsuarioEspecial]

    def check_permission(self, request):
        tipo_usuario = getattr(request.user, 'tipo_usuario', None)
        return tipo_usuario in [1, 3, 4]

    def get(self, request):
        try:
            cargos = Cargo.objects.filter(estadocargo=1).order_by('idcargo')
            serializer = CargoSerializer(cargos, many=True)
            return Response({"count": cargos.count(), "results": serializer.data}, status=200)
        except Exception as e:
            return Response({"error": f"Error al obtener cargos: {str(e)}"}, status=500)

    def post(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para crear cargos"}, status=403)
        try:
            nombre_cargo = request.data.get('nombrecargo')
            if not nombre_cargo:
                return Response({"error": "El campo 'nombrecargo' es requerido"}, status=400)
            if Cargo.objects.filter(nombrecargo=nombre_cargo).exists():
                return Response({"error": "El cargo ya existe"}, status=400)

            cargo = Cargo.objects.create(nombrecargo=nombre_cargo, estadocargo=1)
            serializer = CargoSerializer(cargo)
            return Response({"mensaje": "Cargo creado correctamente", "data": serializer.data}, status=201)
        except Exception as e:
            return Response({"error": f"Error al crear cargo: {str(e)}"}, status=500)

    def put(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para actualizar cargos"}, status=403)
        try:
            cargo_id = request.data.get('idcargo')
            nombre_cargo = request.data.get('nombrecargo')
            if not cargo_id:
                return Response({"error": "El campo 'idcargo' es requerido"}, status=400)
            if not nombre_cargo:
                return Response({"error": "El campo 'nombrecargo' es requerido"}, status=400)

            cargo = Cargo.objects.filter(idcargo=cargo_id).first()
            if not cargo:
                return Response({"error": "Cargo no encontrado"}, status=404)

            cargo.nombrecargo = nombre_cargo
            cargo.save()
            serializer = CargoSerializer(cargo)
            return Response({"mensaje": "Cargo actualizado correctamente", "data": serializer.data}, status=200)
        except Exception as e:
            return Response({"error": f"Error al actualizar cargo: {str(e)}"}, status=500)

    def delete(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para eliminar cargos"}, status=403)
        try:
            cargo_id = request.data.get('idcargo')
            if not cargo_id:
                return Response({"error": "El campo 'idcargo' es requerido"}, status=400)

            cargo = Cargo.objects.filter(idcargo=cargo_id).first()
            if not cargo:
                return Response({"error": "Cargo no encontrado"}, status=404)

            cargo.estadocargo = 0
            cargo.save()
            return Response({"mensaje": "Cargo desactivado correctamente", "cargo_id": cargo.idcargo}, status=200)
        except Exception as e:
            return Response({"error": f"Error al eliminar cargo: {str(e)}"}, status=500)


class DatosNivelView(APIView):
    """CRUD para gestionar Niveles."""
    permission_classes = [IsAuthenticated, IsSuperUserOrAdmin | IsUsuarioEspecial]

    def check_permission(self, request):
        tipo_usuario = getattr(request.user, 'tipo_usuario', None)
        return tipo_usuario in [1, 3, 4]

    def get(self, request):
        try:
            niveles = Niveles.objects.filter(estadonivel=1).order_by('idnivel')
            serializer = NivelesSerializer(niveles, many=True)
            return Response({"count": niveles.count(), "results": serializer.data}, status=200)
        except Exception as e:
            return Response({"error": f"Error al obtener niveles: {str(e)}"}, status=500)

    def post(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para crear niveles"}, status=403)
        try:
            nombre_nivel = request.data.get('nombrenivel')
            if not nombre_nivel:
                return Response({"error": "El campo 'nombrenivel' es requerido"}, status=400)
            if Niveles.objects.filter(nombrenivel=nombre_nivel).exists():
                return Response({"error": "El nivel ya existe"}, status=400)

            nivel = Niveles.objects.create(nombrenivel=nombre_nivel, estadonivel=1)
            serializer = NivelesSerializer(nivel)
            return Response({"mensaje": "Nivel creado correctamente", "data": serializer.data}, status=201)
        except Exception as e:
            return Response({"error": f"Error al crear nivel: {str(e)}"}, status=500)

    def put(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para actualizar niveles"}, status=403)
        try:
            nivel_id = request.data.get('idnivel')
            nombre_nivel = request.data.get('nombrenivel')
            if not nivel_id:
                return Response({"error": "El campo 'idnivel' es requerido"}, status=400)
            if not nombre_nivel:
                return Response({"error": "El campo 'nombrenivel' es requerido"}, status=400)

            nivel = Niveles.objects.filter(idnivel=nivel_id).first()
            if not nivel:
                return Response({"error": "Nivel no encontrado"}, status=404)

            nivel.nombrenivel = nombre_nivel
            nivel.save()
            serializer = NivelesSerializer(nivel)
            return Response({"mensaje": "Nivel actualizado correctamente", "data": serializer.data}, status=200)
        except Exception as e:
            return Response({"error": f"Error al actualizar nivel: {str(e)}"}, status=500)

    def delete(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para eliminar niveles"}, status=403)
        try:
            nivel_id = request.data.get('idnivel')
            if not nivel_id:
                return Response({"error": "El campo 'idnivel' es requerido"}, status=400)

            nivel = Niveles.objects.filter(idnivel=nivel_id).first()
            if not nivel:
                return Response({"error": "Nivel no encontrado"}, status=404)

            nivel.estadonivel = 0
            nivel.save()
            return Response({"mensaje": "Nivel desactivado correctamente", "nivel_id": nivel.idnivel}, status=200)
        except Exception as e:
            return Response({"error": f"Error al eliminar nivel: {str(e)}"}, status=500)


class DatosRegionView(APIView):
    """CRUD para gestionar Regionales."""
    permission_classes = [IsAuthenticated, IsSuperUserOrAdmin | IsUsuarioEspecial]

    def check_permission(self, request):
        tipo_usuario = getattr(request.user, 'tipo_usuario', None)
        return tipo_usuario in [1, 3, 4]

    def get(self, request):
        try:
            regionales = Regional.objects.filter(estadoregional=1).order_by('idregional')
            serializer = RegionalSerializer(regionales, many=True)
            return Response({"count": regionales.count(), "results": serializer.data}, status=200)
        except Exception as e:
            return Response({"error": f"Error al obtener regionales: {str(e)}"}, status=500)

    def post(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para crear regionales"}, status=403)
        try:
            nombre_regional = request.data.get('nombreregional')
            if not nombre_regional:
                return Response({"error": "El campo 'nombreregional' es requerido"}, status=400)
            if Regional.objects.filter(nombreregional=nombre_regional).exists():
                return Response({"error": "La regional ya existe"}, status=400)

            regional = Regional.objects.create(nombreregional=nombre_regional, estadoregional=1)
            serializer = RegionalSerializer(regional)
            return Response({"mensaje": "Regional creada correctamente", "data": serializer.data}, status=201)
        except Exception as e:
            return Response({"error": f"Error al crear regional: {str(e)}"}, status=500)

    def put(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para actualizar regionales"}, status=403)
        try:
            regional_id = request.data.get('idregional')
            nombre_regional = request.data.get('nombreregional')
            if not regional_id:
                return Response({"error": "El campo 'idregional' es requerido"}, status=400)
            if not nombre_regional:
                return Response({"error": "El campo 'nombreregional' es requerido"}, status=400)

            regional = Regional.objects.filter(idregional=regional_id).first()
            if not regional:
                return Response({"error": "Regional no encontrada"}, status=404)

            regional.nombreregional = nombre_regional
            regional.save()
            serializer = RegionalSerializer(regional)
            return Response({"mensaje": "Regional actualizada correctamente", "data": serializer.data}, status=200)
        except Exception as e:
            return Response({"error": f"Error al actualizar regional: {str(e)}"}, status=500)

    def delete(self, request):
        if not self.check_permission(request):
            return Response({"error": "No tiene permisos para eliminar regionales"}, status=403)
        try:
            regional_id = request.data.get('idregional')
            if not regional_id:
                return Response({"error": "El campo 'idregional' es requerido"}, status=400)

            regional = Regional.objects.filter(idregional=regional_id).first()
            if not regional:
                return Response({"error": "Regional no encontrada"}, status=404)

            regional.estadoregional = 0
            regional.save()
            return Response({"mensaje": "Regional desactivada correctamente", "regional_id": regional.idregional}, status=200)
        except Exception as e:
            return Response({"error": f"Error al eliminar regional: {str(e)}"}, status=500)


class RegistrarMasivoView(APIView):
    """
    Vista para registrar multiples usuarios a traves de un archivo CSV UTF-8.

    CSV esperado (separador ;):
    cedula;nombre_completo;correo;telefono;region;nivel;cargo

    - La cedula se usa como usuario y contrasena inicial.
    - tipo_usuario por defecto es 0 (usuario normal).
    """
    permission_classes = [IsAuthenticated, IsSuperUserOrAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def _buscar_cargo(self, nombre_cargo):
        return Cargo.objects.filter(nombrecargo__iexact=nombre_cargo.strip(), estadocargo=1).first()

    def _buscar_nivel(self, nombre_nivel):
        return Niveles.objects.filter(nombrenivel__iexact=nombre_nivel.strip(), estadonivel=1).first()

    def _buscar_regional(self, nombre_regional):
        return Regional.objects.filter(nombreregional__iexact=nombre_regional.strip(), estadoregional=1).first()

    def _get_reader_and_mapeo(self, archivo):
        try:
            contenido = archivo.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                archivo.seek(0)
                contenido = archivo.read().decode('latin-1')
            except Exception:
                return None, None, "No se pudo leer el archivo. Asegurese de que este en formato UTF-8."

        primera_linea = contenido.split('\n')[0]
        delimitador = ';' if ';' in primera_linea else ','

        reader = csv.DictReader(io.StringIO(contenido), delimiter=delimitador)
        if reader.fieldnames is None:
            return None, None, "El archivo CSV esta vacio o no tiene encabezados."

        columnas_normalizadas = {col.strip(): col.strip().lower() for col in reader.fieldnames}
        mapeo_columnas = {
            'cedula': None, 'nombre': None, 'correo': None, 'telefono': None,
            'region': None, 'nivel': None, 'cargo': None
        }

        busquedas = [
            ('cedula', ['cedula', 'cc', 'id']),
            ('telefono', ['telefono', 'numero', 'celular']),
            ('region', ['region', 'regional']),
            ('nivel', ['nivel']),
            ('cargo', ['cargo']),
            ('nombre', ['nombre', 'nombre_completo']),
            ('correo', ['correo', 'email']),
        ]

        for esperada, patrones in busquedas:
            for col_original, col_norm in columnas_normalizadas.items():
                if mapeo_columnas[esperada] is None:
                    if any(p == col_norm or p in col_norm for p in patrones):
                        mapeo_columnas[esperada] = col_original
                        break

        return reader, mapeo_columnas, None

    def post(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({"error": "Se requiere un archivo CSV. Envielo con el campo 'archivo'."}, status=400)
        if not archivo.name.lower().endswith('.csv'):
            return Response({"error": "El archivo debe ser de tipo .csv"}, status=400)

        try:
            reader, mapeo_columnas, error = self._get_reader_and_mapeo(archivo)
            if error:
                return Response({"error": error}, status=400)

            faltantes = [c for c in ['cedula', 'nombre'] if mapeo_columnas[c] is None]
            if faltantes:
                return Response(
                    {"error": f"Faltan columnas requeridas: {', '.join(faltantes)}. Encontradas: {', '.join(reader.fieldnames)}"},
                    status=400
                )

            filas_datos = []
            errores_validacion = []
            cedulas_en_csv = {}

            for num_fila, fila in enumerate(reader, start=2):
                fila_limpia = {k.strip(): (v.strip() if v else '') for k, v in fila.items()}

                cedula = fila_limpia.get(mapeo_columnas.get('cedula', ''), '').strip()
                nombre_completo = fila_limpia.get(mapeo_columnas.get('nombre', ''), '').strip()
                correo = fila_limpia.get(mapeo_columnas.get('correo', ''), '').strip() if mapeo_columnas.get('correo') else ''
                telefono = fila_limpia.get(mapeo_columnas.get('telefono', ''), '').strip() if mapeo_columnas.get('telefono') else ''
                region_nombre = fila_limpia.get(mapeo_columnas.get('region', ''), '').strip() if mapeo_columnas.get('region') else ''
                nivel_nombre = fila_limpia.get(mapeo_columnas.get('nivel', ''), '').strip() if mapeo_columnas.get('nivel') else ''
                cargo_nombre = fila_limpia.get(mapeo_columnas.get('cargo', ''), '').strip() if mapeo_columnas.get('cargo') else ''

                if not cedula or not nombre_completo:
                    continue

                if cedula in cedulas_en_csv:
                    errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Cedula duplicada en el CSV (fila {cedulas_en_csv[cedula]})"})
                    continue
                cedulas_en_csv[cedula] = num_fila

                if Usuario.objects.filter(cedula=cedula).exists():
                    errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Cedula {cedula} ya registrada"})
                    continue

                if Credenciales.objects.filter(usuario=cedula).exists():
                    errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Usuario {cedula} ya existe"})
                    continue

                cargo_obj = None
                if cargo_nombre:
                    cargo_obj = self._buscar_cargo(cargo_nombre)
                    if not cargo_obj:
                        errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Cargo no encontrado: {cargo_nombre}"})
                        continue

                nivel_obj = None
                if nivel_nombre:
                    nivel_obj = self._buscar_nivel(nivel_nombre)
                    if not nivel_obj:
                        errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Nivel no encontrado: {nivel_nombre}"})
                        continue

                regional_obj = None
                if region_nombre:
                    regional_obj = self._buscar_regional(region_nombre)
                    if not regional_obj:
                        errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Regional no encontrada: {region_nombre}"})
                        continue

                filas_datos.append({
                    "num_fila": num_fila,
                    "cedula": cedula,
                    "nombre_completo": nombre_completo,
                    "correo": correo,
                    "telefono": telefono,
                    "cargo_obj": cargo_obj,
                    "nivel_obj": nivel_obj,
                    "regional_obj": regional_obj,
                })

            if errores_validacion:
                return Response({
                    "error": "Validacion fallida. No se creo ningun registro.",
                    "total_errores": len(errores_validacion),
                    "detalles_errores": errores_validacion
                }, status=400)

            if not filas_datos:
                return Response({"error": "El archivo CSV no contiene filas validas para procesar", "total_filas": 0, "creados": 0}, status=400)

            resultados = []
            try:
                with transaction.atomic():
                    for fila_data in filas_datos:
                        usuario_obj = Usuario.objects.create(
                            cedula=fila_data['cedula'],
                            nombre_completo=fila_data['nombre_completo'],
                            correo=fila_data['correo'],
                            telefono=fila_data['telefono'],
                            cargo=fila_data['cargo_obj'],
                            nivel=fila_data['nivel_obj'],
                            regional=fila_data['regional_obj'],
                            estado=1,
                        )
                        cred = Credenciales(
                            usuario=fila_data['cedula'],
                            tipo_usuario=0,
                            usuario_rel=usuario_obj,
                            estado=1,
                        )
                        cred.set_password(fila_data['cedula'])
                        cred.save()

                        resultados.append({
                            "fila": fila_data['num_fila'],
                            "cedula": fila_data['cedula'],
                            "nombre_completo": fila_data['nombre_completo'],
                            "usuario_id": cred.id,
                            "success": True,
                        })

                return Response({
                    "mensaje": f"Todos los {len(resultados)} usuarios fueron registrados exitosamente",
                    "total_creados": len(resultados),
                    "detalles": resultados
                }, status=201)

            except Exception as e:
                return Response({
                    "error": f"Error durante la creacion. Ningun usuario fue creado. Detalles: {str(e)}",
                    "total_intentados": len(filas_datos),
                    "creados": 0
                }, status=500)

        except Exception as e:
            return Response({"error": f"Error al procesar el archivo CSV: {str(e)}"}, status=500)

    def put(self, request):
        """Actualizacion masiva de usuarios existentes a traves de un archivo CSV."""
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({"error": "Se requiere un archivo CSV. Envielo con el campo 'archivo'."}, status=400)
        if not archivo.name.lower().endswith('.csv'):
            return Response({"error": "El archivo debe ser de tipo .csv"}, status=400)

        try:
            reader, mapeo_columnas, error = self._get_reader_and_mapeo(archivo)
            if error:
                return Response({"error": error}, status=400)

            if mapeo_columnas['cedula'] is None:
                return Response(
                    {"error": f"Falta la columna 'cedula'. Columnas encontradas: {', '.join(reader.fieldnames)}"},
                    status=400
                )

            filas_nuevas = []
            filas_existentes = []
            errores_validacion = []
            cedulas_en_csv = {}

            for num_fila, fila in enumerate(reader, start=2):
                fila_limpia = {k.strip(): (v.strip() if v else '') for k, v in fila.items()}

                cedula = fila_limpia.get(mapeo_columnas.get('cedula', ''), '').strip()
                nombre_completo = fila_limpia.get(mapeo_columnas.get('nombre', ''), '').strip()
                correo = fila_limpia.get(mapeo_columnas.get('correo', ''), '').strip() if mapeo_columnas.get('correo') else ''
                telefono = fila_limpia.get(mapeo_columnas.get('telefono', ''), '').strip() if mapeo_columnas.get('telefono') else ''
                region_nombre = fila_limpia.get(mapeo_columnas.get('region', ''), '').strip() if mapeo_columnas.get('region') else ''
                nivel_nombre = fila_limpia.get(mapeo_columnas.get('nivel', ''), '').strip() if mapeo_columnas.get('nivel') else ''
                cargo_nombre = fila_limpia.get(mapeo_columnas.get('cargo', ''), '').strip() if mapeo_columnas.get('cargo') else ''

                if not cedula:
                    continue

                if cedula in cedulas_en_csv:
                    errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Cedula duplicada (fila {cedulas_en_csv[cedula]})"})
                    continue
                cedulas_en_csv[cedula] = num_fila

                usuario_existente = Usuario.objects.filter(cedula=cedula).first()
                es_nuevo = usuario_existente is None

                if es_nuevo and not nombre_completo:
                    errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": "Nombre requerido para nuevo usuario"})
                    continue

                cargo_obj = None
                if cargo_nombre:
                    cargo_obj = self._buscar_cargo(cargo_nombre)
                    if not cargo_obj:
                        errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Cargo '{cargo_nombre}' no encontrado"})
                        continue

                nivel_obj = None
                if nivel_nombre:
                    nivel_obj = self._buscar_nivel(nivel_nombre)
                    if not nivel_obj:
                        errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Nivel '{nivel_nombre}' no encontrado"})
                        continue

                regional_obj = None
                if region_nombre:
                    regional_obj = self._buscar_regional(region_nombre)
                    if not regional_obj:
                        errores_validacion.append({"fila": num_fila, "cedula": cedula, "error": f"Regional '{region_nombre}' no encontrada"})
                        continue

                fila_obj = {
                    "num_fila": num_fila,
                    "cedula": cedula,
                    "nombre_completo": nombre_completo,
                    "correo": correo,
                    "telefono": telefono,
                    "cargo_obj": cargo_obj,
                    "nivel_obj": nivel_obj,
                    "regional_obj": regional_obj,
                }

                if es_nuevo:
                    filas_nuevas.append(fila_obj)
                else:
                    fila_obj['usuario_obj'] = usuario_existente
                    filas_existentes.append(fila_obj)

            if errores_validacion:
                return Response({
                    "error": "Validacion fallida. No se actualizo ningun registro.",
                    "total_errores": len(errores_validacion),
                    "detalles_errores": errores_validacion
                }, status=400)

            if not filas_nuevas and not filas_existentes:
                return Response({"error": "El archivo CSV no contiene filas validas", "total_filas": 0, "procesados": 0}, status=400)

            resultados = []
            try:
                with transaction.atomic():
                    for fila_data in filas_nuevas:
                        usuario_obj = Usuario.objects.create(
                            cedula=fila_data['cedula'],
                            nombre_completo=fila_data['nombre_completo'],
                            correo=fila_data['correo'],
                            telefono=fila_data['telefono'],
                            cargo=fila_data['cargo_obj'],
                            nivel=fila_data['nivel_obj'],
                            regional=fila_data['regional_obj'],
                            estado=1,
                        )
                        cred = Credenciales(
                            usuario=fila_data['cedula'],
                            tipo_usuario=0,
                            usuario_rel=usuario_obj,
                            estado=1,
                        )
                        cred.set_password(fila_data['cedula'])
                        cred.save()

                        resultados.append({
                            "fila": fila_data['num_fila'],
                            "cedula": fila_data['cedula'],
                            "nombre_completo": fila_data['nombre_completo'],
                            "accion": "CREADO",
                            "success": True,
                        })

                    for fila_data in filas_existentes:
                        usuario_obj = fila_data['usuario_obj']
                        cambios = []

                        if fila_data['nombre_completo'] and fila_data['nombre_completo'] != usuario_obj.nombre_completo:
                            usuario_obj.nombre_completo = fila_data['nombre_completo']
                            cambios.append('nombre_completo')
                        if fila_data['correo'] and fila_data['correo'] != (usuario_obj.correo or ''):
                            usuario_obj.correo = fila_data['correo']
                            cambios.append('correo')
                        if fila_data['telefono'] and fila_data['telefono'] != (usuario_obj.telefono or ''):
                            usuario_obj.telefono = fila_data['telefono']
                            cambios.append('telefono')
                        if fila_data['cargo_obj'] and fila_data['cargo_obj'] != usuario_obj.cargo:
                            usuario_obj.cargo = fila_data['cargo_obj']
                            cambios.append('cargo')
                        if fila_data['nivel_obj'] and fila_data['nivel_obj'] != usuario_obj.nivel:
                            usuario_obj.nivel = fila_data['nivel_obj']
                            cambios.append('nivel')
                        if fila_data['regional_obj'] and fila_data['regional_obj'] != usuario_obj.regional:
                            usuario_obj.regional = fila_data['regional_obj']
                            cambios.append('regional')

                        usuario_obj.save()

                        resultados.append({
                            "fila": fila_data['num_fila'],
                            "cedula": fila_data['cedula'],
                            "nombre_completo": usuario_obj.nombre_completo,
                            "campos_actualizados": cambios,
                            "accion": "ACTUALIZADO",
                            "success": True,
                        })

                creados = len([r for r in resultados if r.get('accion') == 'CREADO'])
                actualizados = len([r for r in resultados if r.get('accion') == 'ACTUALIZADO'])

                return Response({
                    "mensaje": f"Procesamiento completado: {creados} creados, {actualizados} actualizados",
                    "total_creados": creados,
                    "total_actualizados": actualizados,
                    "total_procesados": len(resultados),
                    "detalles": resultados
                }, status=200)

            except Exception as e:
                return Response({
                    "error": f"Error durante el procesamiento. Ningun usuario fue modificado o creado. Detalles: {str(e)}",
                    "total_intentados": len(filas_nuevas) + len(filas_existentes),
                    "creados": 0,
                    "actualizados": 0
                }, status=500)

        except Exception as e:
            return Response({"error": f"Error al procesar el archivo CSV: {str(e)}"}, status=500)

    def get(self, request):
        """Devuelve la plantilla CSV de ejemplo como descarga."""
        template_path = os.path.join(
            os.path.dirname(__file__), 'templates', 'Registrar usuarios - ejemplo.csv'
        )
        if os.path.exists(template_path):
            return FileResponse(
                open(template_path, 'rb'),
                as_attachment=True,
                filename='plantilla_registro_masivo.csv'
            )
        return Response({"error": "Plantilla no encontrada"}, status=404)


class ReporteUsuariosView(APIView):
    """Vista para generar reporte de usuarios en formato Excel."""
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsAdminUser]

    def get(self, request):
        try:
            usuarios = (
                Usuario.objects
                .select_related('cargo', 'nivel', 'regional')
                .exclude(estado=3)
                .order_by('id')
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "Usuarios"

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            columnas = ['Cedula', 'Nombre Completo', 'Correo', 'Telefono', 'Sede', 'Regional', 'Nivel', 'Cargo', 'Estado']

            for col_num, col_titulo in enumerate(columnas, start=1):
                celda = ws.cell(row=1, column=col_num)
                celda.value = col_titulo
                celda.fill = header_fill
                celda.font = header_font
                celda.alignment = header_alignment
                celda.border = border

            for row_num, usuario in enumerate(usuarios, start=2):
                estado_texto = 'Activo' if usuario.estado == 1 else ('Inactivo' if usuario.estado == 0 else 'N/A')
                datos = [
                    usuario.cedula or '',
                    usuario.nombre_completo or '',
                    usuario.correo or '',
                    usuario.telefono or '',
                    usuario.sede or '',
                    usuario.regional.nombreregional if usuario.regional else '',
                    usuario.nivel.nombrenivel if usuario.nivel else '',
                    usuario.cargo.nombrecargo if usuario.cargo else '',
                    estado_texto,
                ]

                for col_num, valor in enumerate(datos, start=1):
                    celda = ws.cell(row=row_num, column=col_num)
                    celda.value = valor
                    celda.alignment = center_alignment
                    celda.border = border

            anchos = [15, 30, 25, 15, 20, 20, 20, 20, 10]
            for col_num, ancho in enumerate(anchos, start=1):
                ws.column_dimensions[chr(64 + col_num)].width = ancho

            tab = Table(displayName="TablaUsuarios", ref=f"A1:I{max(2, ws.max_row)}")
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            archivo_excel = BytesIO()
            wb.save(archivo_excel)
            archivo_excel.seek(0)

            return FileResponse(
                archivo_excel,
                as_attachment=True,
                filename='Reporte_Usuarios.xlsx',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        except Exception as e:
            return Response({"error": f"Error al generar el reporte: {str(e)}"}, status=500)
